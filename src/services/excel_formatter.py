"""
Excel formatting service for professional-looking spreadsheets.

This module provides utilities to export DataFrames to Excel files
with consistent styling, colors, and formatting.
"""

from pathlib import Path
from typing import Optional, List, Dict, Any
import pandas as pd
import logging

from openpyxl import Workbook
from openpyxl.styles import (
    Font, Fill, PatternFill, Alignment, Border, Side, NamedStyle
)
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
import unicodedata

logger = logging.getLogger(__name__)


class ExcelTheme:
    """Color theme for Excel formatting."""
    
    # Header colors
    HEADER_BG = "2F5496"  # Dark blue
    HEADER_FG = "FFFFFF"  # White text
    
    # Alternating row colors
    ROW_EVEN = "D6E3F8"  # Light blue
    ROW_ODD = "FFFFFF"   # White
    
    # Special row colors
    SUMMARY_BG = "FFF2CC"  # Light yellow for GERAL/MédiaTodosDias rows
    SUMMARY_FG = "000000"  # Black text
    
    # Border color
    BORDER_COLOR = "B4C6E7"  # Light blue border


class ExcelStyles:
    """Pre-defined styles for Excel formatting."""
    
    @staticmethod
    def get_header_font() -> Font:
        """Bold white font for headers."""
        return Font(bold=True, color=ExcelTheme.HEADER_FG, size=11)
    
    @staticmethod
    def get_header_fill() -> PatternFill:
        """Dark blue background for headers."""
        return PatternFill(
            start_color=ExcelTheme.HEADER_BG,
            end_color=ExcelTheme.HEADER_BG,
            fill_type="solid"
        )
    
    @staticmethod
    def get_header_alignment() -> Alignment:
        """Center alignment for headers."""
        return Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    @staticmethod
    def get_data_alignment() -> Alignment:
        """Default alignment for data cells."""
        return Alignment(horizontal="center", vertical="center")
    
    @staticmethod
    def get_even_row_fill() -> PatternFill:
        """Light blue background for even rows."""
        return PatternFill(
            start_color=ExcelTheme.ROW_EVEN,
            end_color=ExcelTheme.ROW_EVEN,
            fill_type="solid"
        )
    
    @staticmethod
    def get_odd_row_fill() -> PatternFill:
        """White background for odd rows."""
        return PatternFill(
            start_color=ExcelTheme.ROW_ODD,
            end_color=ExcelTheme.ROW_ODD,
            fill_type="solid"
        )

    @staticmethod
    def get_team_fill() -> PatternFill:
        """Fill for team-based zebra (light blue #DCE6F1)."""
        return PatternFill(
            start_color="DCE6F1",
            end_color="DCE6F1",
            fill_type="solid",
        )

    @staticmethod
    def get_no_fill() -> PatternFill:
        """Represents no fill for a cell/row."""
        return PatternFill(fill_type=None)
    
    @staticmethod
    def get_summary_fill() -> PatternFill:
        """Yellow background for summary rows."""
        return PatternFill(
            start_color=ExcelTheme.SUMMARY_BG,
            end_color=ExcelTheme.SUMMARY_BG,
            fill_type="solid"
        )
    
    @staticmethod
    def get_summary_font() -> Font:
        """Bold font for summary rows."""
        return Font(bold=True, color=ExcelTheme.SUMMARY_FG, size=11)
    
    @staticmethod
    def get_thin_border() -> Border:
        """Thin border for cells."""
        side = Side(style="thin", color=ExcelTheme.BORDER_COLOR)
        return Border(left=side, right=side, top=side, bottom=side)
    
    @staticmethod
    def get_number_format() -> str:
        """Number format for decimal values."""
        return "#,##0.00"
    
    @staticmethod
    def get_integer_format() -> str:
        """Number format for integer values."""
        return "#,##0"


class ExcelFormatter:
    """
    Service for exporting DataFrames to professionally formatted Excel files.
    
    Features:
    - Styled headers with colors
    - Alternating row colors (zebra striping)
    - Auto-sized columns
    - Highlighted summary rows (GERAL)
    - Proper number formatting
    """
    
    def __init__(self):
        """Initialize the Excel formatter."""
        self._styles = ExcelStyles()
    
    def export(
        self,
        df: pd.DataFrame,
        path: Path = None,
        sheet_name: str = "Dados",
        summary_identifier: str = "GERAL",
        freeze_header: bool = True,
        add_goals_table: bool = None,
        worksheet: Worksheet = None
    ) -> bool:
        """
        Export DataFrame to a formatted Excel file.
        
        Args:
            df: DataFrame to export
            path: Output file path
            sheet_name: Name for the worksheet
            summary_identifier: Text that identifies summary rows (e.g., "GERAL")
            freeze_header: Whether to freeze the header row
            
        Returns:
            True if export was successful, False otherwise
        """
        try:
            if worksheet is not None:
                ws = worksheet
            else:
                logger.info(f"Exporting formatted Excel to: {path}")
                from openpyxl import Workbook
                wb = Workbook()
                ws = wb.active
                ws.title = sheet_name

            # Write data
            self._write_data(ws, df)

            # Apply formatting
            self._format_header(ws, len(df.columns))
            self._format_data_rows(ws, df, summary_identifier)
            self._auto_size_columns(ws, df)

            # Adiciona tabela de metas apenas se for planilha de médias
            if add_goals_table is None:
                add_goals_table = any(
                    s in ws.title.lower() for s in ["médias", "medias", "averages"]
                )
            if add_goals_table:
                self._add_goals_table(ws, df)

            # Freeze header row
            if freeze_header:
                ws.freeze_panes = "A2"

            if worksheet is None and path is not None:
                wb.save(path)
                logger.info(f"Excel file saved successfully: {path}")
            return True
        except Exception as e:
            logger.error(f"Failed to export Excel file: {e}")
            return False

    def _add_goals_table(self, ws: Worksheet, df: pd.DataFrame) -> None:
        """Adiciona a tabela de metas ao lado da tabela de médias como legenda, com colunas autoajustadas."""
        metas = [
            ["Métrica", "Meta Produtivo", "Meta Improdutivo"],
            ["Media_TempExe", "<=50 min", "<=20 min"],
            ["Media_InterReg", "<=60 min", "<=60 min"],
            ["Utilização", ">=85% da Media_Jornada", ">=85% da Media_Jornada"],
            ["Retorno a base", "<=40 min", "<=40 min"],
            ["Media_TempPrepEquipe", "<=10 min", "<=10 min"],
            ["Media_SemOrdemJornada", "<=10 min", "<=10 min"],
            ["Media_AtrasLogin", "<=8", "<=8"],
            ["qtd_ordem", ">=5", ">=5"],
        ]
        start_col = len(df.columns) + 3
        col_widths = [max(len(str(row[c])) for row in metas) for c in range(3)]
        for row_idx, row in enumerate(metas, 1):
            for col_idx, value in enumerate(row, 0):
                cell = ws.cell(row=row_idx, column=start_col + col_idx, value=value)
                cell.font = Font(bold=True) if row_idx == 1 else Font()
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = self._styles.get_thin_border()
                if row_idx == 1:
                    cell.fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
                else:
                    cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        # Ajusta largura das colunas da tabela de metas
        for col_idx, width in enumerate(col_widths, 0):
            col_letter = get_column_letter(start_col + col_idx)
            ws.column_dimensions[col_letter].width = width + 2
    
    def _write_data(self, ws: Worksheet, df: pd.DataFrame) -> None:
        """Write DataFrame data to worksheet, inserting separator rows between teams for 'deslocamento_calculado'."""
        # Write header
        for col_idx, column in enumerate(df.columns, 1):
            ws.cell(row=1, column=col_idx, value=column)

        # Detect team column
        equipe_col = None
        for col in df.columns:
            if "equipe" in col.lower():
                equipe_col = col
                break

        prev_team = None
        excel_row = 2
        for row_idx in range(len(df)):
            # Inserir linha de separação se mudar de equipe (apenas para deslocamento_calculado)
            if ws.title.lower() == "deslocamento_calculado" and equipe_col:
                current_team = df.iloc[row_idx][equipe_col]
                if prev_team is not None and current_team != prev_team:
                    for col_idx in range(1, len(df.columns) + 1):
                        cell = ws.cell(row=excel_row, column=col_idx)
                        cell.value = None
                    excel_row += 1
                prev_team = current_team
            # Escrever dados normalmente
            for col_idx, value in enumerate(df.iloc[row_idx], 1):
                cell = ws.cell(row=excel_row, column=col_idx)
                if pd.isna(value):
                    cell.value = ""
                else:
                    cell.value = value
            excel_row += 1
    
    def _format_header(self, ws: Worksheet, num_cols: int) -> None:
        """Apply formatting to header row."""
        header_font = self._styles.get_header_font()
        header_fill = self._styles.get_header_fill()
        header_alignment = self._styles.get_header_alignment()
        border = self._styles.get_thin_border()
        
        for col_idx in range(1, num_cols + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # Set header row height
        ws.row_dimensions[1].height = 30
    
    def _format_data_rows(
        self,
        ws: Worksheet,
        df: pd.DataFrame,
        summary_identifier: str
    ) -> None:
        """Apply formatting to data rows, with visual separation between teams for 'deslocamento_calculado'.

        Implements:
        - Zebra fill toggling by `Equipe` changes (entire row fill alternates between #DCE6F1 and no fill).
        - Font color toggling by `Data Referência` changes (entire row font alternates between #CC3300 and black).
        - Both toggles are independent and compare each row with the previous data row.
        - Autofilter is enabled for the full table after formatting.
        """
        even_fill = self._styles.get_even_row_fill()
        odd_fill = self._styles.get_odd_row_fill()
        summary_fill = self._styles.get_summary_fill()
        summary_font = self._styles.get_summary_font()
        team_fill_solid = self._styles.get_team_fill()
        no_fill = self._styles.get_no_fill()
        data_alignment = self._styles.get_data_alignment()
        border = self._styles.get_thin_border()
        number_format = self._styles.get_number_format()
        integer_format = self._styles.get_integer_format()

        num_cols = len(df.columns)

        # Identify summary rows (keep existing heuristics)
        summary_row_indices = set()
        if "Data" in df.columns:
            for row_idx, value in enumerate(df["Data"]):
                if str(value) == summary_identifier:
                    summary_row_indices.add(row_idx)

        first_col = df.iloc[:, 0]
        for row_idx, value in enumerate(first_col):
            if "MédiaTodosDias" in str(value):
                summary_row_indices.add(row_idx)

        # Detect team column (tolerant)
        equipe_col = None
        for col in df.columns:
            if "equipe" in col.lower():
                equipe_col = col
                break

        # Detect date/reference column (tolerant to accents/spaces)
        def _norm(s: str) -> str:
            s2 = unicodedata.normalize("NFKD", str(s))
            s2 = s2.encode("ASCII", "ignore").decode()
            return s2.lower().replace(" ", "").replace("_", "")

        date_col = None
        for col in df.columns:
            n = _norm(col)
            if "data" in n and ("refer" in n or "referencia" in n or n == "data"):
                date_col = col
                break
        # fallback: any column named exactly 'data'
        if date_col is None:
            for col in df.columns:
                if _norm(col) == "data":
                    date_col = col
                    break

        # Metas para formatação condicional (unchanged)
        metas_cond = {
            "Media_TempExe": {"produtivo": 50, "improdutivo": 20, "op": "le"},
            "Media_InterReg": {"produtivo": 60, "improdutivo": 60, "op": "le"},
            "Utilização": {"produtivo": 85, "improdutivo": 85, "op": "ge"},
            **({} if ws.title.lower() in ["dados calculados", "deslocamento_calculado"] else {"Retorno a base": {"produtivo": 40, "improdutivo": 40, "op": "le"}}),
            "Media_TempPrepEquipe": {"produtivo": 10, "improdutivo": 10, "op": "le"},
            "Media_SemOrdemJornada": {"produtivo": 10, "improdutivo": 10, "op": "le"},
            "Media_AtrasLogin": {"produtivo": 8, "improdutivo": 8, "op": "le"},
            "qtd_ordem": {"produtivo": 5, "improdutivo": 5, "op": "ge"},
        }
        fill_alert = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # vermelho claro
        font_alert = Font(color="9C0006")

        # State for toggles (independent)
        prev_team_toggle = None
        team_toggle = False
        prev_date_toggle = None
        date_toggle = False

        # State for separator insertion (kept as before)
        prev_team_sep = None

        excel_row = 2
        for row_idx in range(len(df)):
            # Separator row for deslocamento_calculado (unchanged behaviour)
            team_sep = False
            if ws.title.lower() == "deslocamento_calculado" and equipe_col:
                current_team_sep = df.iloc[row_idx][equipe_col]
                if prev_team_sep is not None and current_team_sep != prev_team_sep:
                    for col_idx in range(1, num_cols + 1):
                        cell = ws.cell(row=excel_row, column=col_idx)
                        cell.fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
                        cell.border = border
                    excel_row += 1
                    team_sep = True
                prev_team_sep = current_team_sep

            # Update toggles based on equipe and date changes (compare with previous data row)
            if equipe_col:
                current_team = df.iloc[row_idx][equipe_col]
                if prev_team_toggle is not None and current_team != prev_team_toggle:
                    team_toggle = not team_toggle
                prev_team_toggle = current_team

            if date_col:
                current_date = df.iloc[row_idx][date_col]
                if prev_date_toggle is not None and current_date != prev_date_toggle:
                    date_toggle = not date_toggle
                prev_date_toggle = current_date

            # Determine date-based font color
            date_color = "CC3300" if date_toggle else "000000"

            # Determine if this is a summary row
            is_summary = row_idx in summary_row_indices

            # Determine row fill: summary overrides, otherwise team-based (fallback to even/odd if no equipe column)
            if is_summary:
                row_fill = summary_fill
            else:
                if equipe_col:
                    row_fill = team_fill_solid if team_toggle else no_fill
                else:
                    row_fill = even_fill if row_idx % 2 == 0 else odd_fill

            # Determine base font: apply date color while preserving bold for summaries
            if is_summary:
                base_font = Font(bold=True, color=date_color, size=11)
            else:
                base_font = Font(size=11, color=date_color)

            # Apply formatting to the row cells
            for col_idx in range(1, num_cols + 1):
                cell = ws.cell(row=excel_row, column=col_idx)
                cell.fill = row_fill
                cell.font = base_font
                cell.alignment = data_alignment
                cell.border = border

                # Apply number formatting
                col_name = df.columns[col_idx - 1]
                if self._is_numeric_column(col_name, df):
                    if "qtd" in col_name.lower():
                        cell.number_format = integer_format
                    else:
                        cell.number_format = number_format

                # Apply metas conditional highlighting (overrides font/fill when triggered)
                for meta_key in metas_cond:
                    if meta_key.lower() in col_name.lower():
                        tipo = "produtivo" if "produt" in ws.title.lower() else "improdutivo"
                        meta = metas_cond[meta_key][tipo]
                        op = metas_cond[meta_key]["op"]
                        try:
                            valor = float(cell.value)
                            if op == "le" and valor > meta:
                                cell.fill = fill_alert
                                cell.font = font_alert
                            elif op == "ge" and valor < meta:
                                cell.fill = fill_alert
                                cell.font = font_alert
                        except Exception:
                            pass

            excel_row += 1

        # Enable autofilter for the full header -> last used row range
        try:
            last_row = ws.max_row
            ws.auto_filter.ref = f"A1:{get_column_letter(num_cols)}{last_row}"
        except Exception:
            # If auto-filter fails for any reason, do not break the export
            pass
    
    def _is_numeric_column(self, col_name: str, df: pd.DataFrame) -> bool:
        """Check if a column should have numeric formatting."""
        if col_name not in df.columns:
            return False
        
        # Check if column is numeric
        dtype = df[col_name].dtype
        return pd.api.types.is_numeric_dtype(dtype)
    
    def _auto_size_columns(self, ws: Worksheet, df: pd.DataFrame) -> None:
        """Auto-size column widths based on content."""
        for col_idx, column in enumerate(df.columns, 1):
            # Calculate max width
            max_length = len(str(column))
            
            for row_idx in range(len(df)):
                cell_value = df.iloc[row_idx, col_idx - 1]
                if pd.notna(cell_value):
                    cell_length = len(str(cell_value))
                    max_length = max(max_length, cell_length)
            
            # Add padding and set width
            adjusted_width = min(max_length + 3, 50)  # Cap at 50 chars
            adjusted_width = max(adjusted_width, 10)   # Minimum 10 chars
            
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = adjusted_width


# Singleton instance
_formatter_instance: Optional[ExcelFormatter] = None


def get_excel_formatter() -> ExcelFormatter:
    """Get or create the Excel formatter singleton."""
    global _formatter_instance
    if _formatter_instance is None:
        _formatter_instance = ExcelFormatter()
    return _formatter_instance
